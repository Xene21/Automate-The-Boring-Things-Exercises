import openpyxl
from openpyxl.utils import get_column_letter
def selectWorkbook():
    chosenBook = input("What excel file would you like to insert blank rows in: ")
    chosenBook += (".xlsx")
    chosenWorkBook = openpyxl.load_workbook(chosenBook)
    return chosenWorkBook


    
def rowInsert(rowsinsertPoint,blanksAdded,workBook):
    
    activeSheet = workBook.active
    newSheet = workBook.create_sheet(title= "BR Sheet")
    for x in range(1,rowsinsertPoint):
            count = 1
            while True:
                
                columnLetter = get_column_letter(count)
                
                if activeSheet[columnLetter + str(x)].value == None:
                     break
                newSheet[columnLetter + str(x)] = activeSheet[columnLetter + str(x)].value 
                count += 1
    for y in range(rowsinsertPoint,activeSheet.max_row+1):
         num = 1
         while True:
            columnLetter = get_column_letter(num)
            if activeSheet[columnLetter + str(y)].value == None:
                     break
            newSheet[columnLetter + str(y+blanksAdded)] = activeSheet[columnLetter + str(y)].value
            num += 1
            
            
    workBook.save("nameOfXenians.xlsx")
    print("Saved")
        


    
blankRowsPassed = int(input("How many blank Rows do you want to pass: "))
pointInsert = int(input("which point would you like to insert the blank rows(Number): "))

bookToBePassed = selectWorkbook()
rowInsert(pointInsert,blankRowsPassed,bookToBePassed)
