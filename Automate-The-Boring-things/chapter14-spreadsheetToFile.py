import openpyxl
from openpyxl.utils import get_column_letter
def loadWorkbook(workbook):
    fileWorkBook = openpyxl.load_workbook(workbook)
    activeSheet = fileWorkBook.active
    return activeSheet
def cellToText(sheet):
    for x in range(1,sheet.max_column+1):
        columnLetter = get_column_letter(x)
        nameOffile = input("What's the name of the file: ")
        file = open(nameOffile +".txt","w")
        for y in range(1,sheet.max_row+1):
            file.write(sheet[columnLetter + str(y)].value)
            
        file.close()
workbookToUse = input("Input Workbook: ")
derivedWorkBook = loadWorkbook(workbookToUse +".xlsx")
cellToText(derivedWorkBook)
#Doesn't run past the first iteration... Future me it's your job to fix it




