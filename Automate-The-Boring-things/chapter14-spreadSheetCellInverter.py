import openpyxl
from openpyxl.utils import get_column_letter

def selectWorkbook():
    chosenBook = input("What excel file would you like to invert: ")
    chosenBook += (".xlsx")
    chosenWorkBook = openpyxl.load_workbook(chosenBook)
    return chosenWorkBook

def cellInverter(workbook):
    activeSheet = workbook.active
    workbook.create_sheet(title= "invertedCell")
    invertedSheet = workbook["invertedCell"]

    for x in range(1,activeSheet.max_column+1):
        for y in range(1,activeSheet.max_row+1):
            activeGetColumn = get_column_letter(x)
            invertedGetColumn = get_column_letter(y)

            invertedSheet[invertedGetColumn + str(x)] = activeSheet[activeGetColumn + str(y)].value

    workbook.save("InvertedCell.xlsx")
    print("INVERTED".rjust(40))



bookToBePassed = selectWorkbook()
cellInverter(bookToBePassed)