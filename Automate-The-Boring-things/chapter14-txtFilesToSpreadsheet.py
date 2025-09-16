import openpyxl
from openpyxl.utils import get_column_letter
def createWorkBook():
    fileWorkbook = openpyxl.Workbook()
    fileWorkbook.create_sheet(title= "Files")
    del fileWorkbook["Sheet"]
    return fileWorkbook

def txtLines(numberOfFiles: int):
    dictTxt = {}
    for n in range(numberOfFiles):
        nameOfFile = input("What's the name of the file: ")
        file = open(nameOfFile + ".txt","r")
        lines = file.readlines()
        dictTxt[nameOfFile] = lines
    
    return dictTxt    
def txtToExcel(txtContent: dict,workbook):
    activeSheet = workbook.active
    i = 0
    for v in txtContent.values():
        i += 1
        columnLetter = get_column_letter(i)
        activeSheet.column_dimensions[columnLetter].width = 20
        for x in range(len(v)):
            activeSheet[columnLetter + str(x + 1)] = v[x]
    print("DONE WITH CONVERSION".center(59))
    workbook.save("FileToTxt.xlsx")
        

fileNumber = int(input("How many txt Files would you like to input(number): "))
passLinesDict = txtLines(fileNumber)
workBookFile = createWorkBook()
txtToExcel(passLinesDict,workBookFile)
