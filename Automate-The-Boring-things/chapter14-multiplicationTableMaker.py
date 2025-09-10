import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font
def openingWorkbook():
    multiplicationWorkbook = openpyxl.Workbook()
    multiplicationWorkbook.create_sheet(title="multiplication table")
    del multiplicationWorkbook["Sheet"]
    return multiplicationWorkbook
def multiplicationTable():
    multiplicationNumber = int(input("Multiplication Table for (Type a number): "))
    return multiplicationNumber
def populatingSheet(workbook):
    numberMultiply = multiplicationTable()
    multiplicationSheet = workbook["multiplication table"]
    #Adding bold to the cells 
    
    for x in range(numberMultiply+1):
        '''number gotten from multiplication Function needs to be incremented by 1... 6 given python
        goes from 0 to 6 and since since the zero index hass been omitted only five values would be given''' 
        if x == 0:
            continue
        columnLetter = get_column_letter(x+1)
        multiplicationSheet["A"+str(x+1)] = x #Accessing each cell using the A column and specified number of rows
        multiplicationSheet["A"+str(x+1)].font = Font(bold=True)
        multiplicationSheet[columnLetter + "1"] = x
        multiplicationSheet[columnLetter + "1"].font = Font(bold=True) #Bold added
        #Cells have a font attr that can be assigned a font object...
        for x in range(1,numberMultiply+1):
            
            mColumnLetter = get_column_letter(x+1)
            for y in range(1,numberMultiply+1):
                if multiplicationSheet['A' + str(y+1)].value == None or multiplicationSheet[mColumnLetter + "1"].value == None:
                    continue
                '''The variables below will contain None values in them a considerable number of times... based on how both loops run
                Eventually a section were no None types exist will be found... run the for loop till then with continue'''

                multiplicationSheet[mColumnLetter + str(y+1)] = multiplicationSheet[mColumnLetter + "1"].value * multiplicationSheet['A' + str(y+1)].value
        # Attempting to multiply... column by all rows(till completion)



    workbook.save("multiplicationTable.xlsx")
    print("Done Populating spreadSheet")


workBookMade = openingWorkbook()
populatingSheet(workBookMade)



